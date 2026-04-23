import os
import json
import logging
from datetime import datetime, time
from apscheduler.schedulers.background import BackgroundScheduler
from telegram import Bot, Update
from telegram.ext import Application, CommandHandler, ContextTypes
import pandas as pd
from openpyxl import load_workbook

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Load configuration
CONFIG_FILE = 'config.json'

def load_config():
    """Load bot configuration from config.json"""
    if not os.path.exists(CONFIG_FILE):
        raise FileNotFoundError(f"{CONFIG_FILE} not found. Please create it with your settings.")
    
    with open(CONFIG_FILE, 'r') as f:
        return json.load(f)

def load_schedule_from_excel(excel_file):
    """Load the cleaning schedule from Excel file"""
    wb = load_workbook(excel_file)
    sheet = wb.active
    
    schedule = {}
    
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[2]:  # Date and Members columns
            try:
                # Parse date
                date_obj = pd.to_datetime(row[0]).date()
                date_str = date_obj.strftime('%Y-%m-%d')
                
                # Parse members
                members = [m.strip() for m in str(row[2]).split(',')]
                team = row[1]
                
                schedule[date_str] = {
                    'members': members,
                    'team': team
                }
            except Exception as e:
                logger.error(f"Error parsing row {row}: {e}")
                continue
    
    return schedule

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Start command - register user"""
    config = load_config()
    user_name = update.effective_user.first_name
    user_id = update.effective_user.id
    
    # Store user info
    if 'users' not in context.bot_data:
        context.bot_data['users'] = {}
    
    context.bot_data['users'][user_id] = {
        'name': user_name,
        'chat_id': update.effective_chat.id
    }
    
    await update.message.reply_text(
        f"🤖 Hello {user_name}! I'm the Coffee Cleaning Reminder Bot.\n\n"
        "I'll remind you when it's your shift to clean the coffee machine at 4:30 PM.\n\n"
        "Commands:\n"
        "/schedule - View your upcoming cleaning shifts\n"
        "/help - Get help"
    )

async def schedule_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Show user's upcoming cleaning shifts"""
    config = load_config()
    schedule = load_schedule_from_excel(config['excel_file'])
    
    user_name = update.effective_user.first_name
    
    upcoming = []
    for date_str in sorted(schedule.keys()):
        date_obj = pd.to_datetime(date_str).date()
        if date_obj >= datetime.now().date():
            members = schedule[date_str]['members']
            if user_name in members or any(user_name.lower() in m.lower() for m in members):
                team = schedule[date_str]['team']
                upcoming.append(f"📅 {date_str} - {team}")
    
    if upcoming:
        message = f"Your upcoming cleaning shifts:\n\n" + "\n".join(upcoming)
    else:
        message = "No upcoming cleaning shifts found for you."
    
    await update.message.reply_text(message)

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Help command"""
    await update.message.reply_text(
        "☕ Coffee Cleaning Reminder Bot\n\n"
        "This bot sends reminders at 4:30 PM on your assigned cleaning days.\n\n"
        "Make sure to:\n"
        "1️⃣ Start the bot with /start\n"
        "2️⃣ Your first name must match the schedule\n"
        "3️⃣ The bot runs 24/7 to send reminders\n\n"
        "Commands:\n"
        "/schedule - View your shifts\n"
        "/help - Show this message"
    )

async def send_daily_reminders(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Send reminders ONLY to members scheduled for today at 4:30 PM"""
    config = load_config()
    schedule = load_schedule_from_excel(config['excel_file'])
    
    today = datetime.now().date()
    today_str = today.strftime('%Y-%m-%d')
    
    # Only send if today is in the schedule
    if today_str in schedule:
        shift_info = schedule[today_str]
        scheduled_members = shift_info['members']
        team = shift_info['team']
        
        message = (
            f"🚨 Coffee Cleaning Reminder!\n\n"
            f"Dear {' & '.join(scheduled_members)},\n\n"
            f"It's your turn to clean the coffee machine today! ☕\n"
            f"Team: {team}\n\n"
            f"Please clean it before end of shift.\n"
            f"Thank you! 😊"
        )
        
        # Send ONLY to members scheduled for today
        if 'users' in context.bot_data:
            for user_id, user_info in context.bot_data['users'].items():
                user_name = user_info['name']
                
                # Check if this user is scheduled for today
                if any(user_name.lower() in member.lower() or member.lower() in user_name.lower() for member in scheduled_members):
                    try:
                        await context.bot.send_message(
                            chat_id=user_info['chat_id'],
                            text=message
                        )
                        logger.info(f"Reminder sent to {user_name} for {today_str}")
                    except Exception as e:
                        logger.error(f"Failed to send message to {user_id}: {e}")
        
        logger.info(f"Reminders sent for {today_str}: {scheduled_members}")
    else:
        logger.info(f"No cleaning scheduled for {today_str} - no reminders sent")

def main() -> None:
    """Start the bot"""
    config = load_config()
    
    # Create application
    app = Application.builder().token(config['telegram_token']).build()
    
    # Add handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("schedule", schedule_command))
    app.add_handler(CommandHandler("help", help_command))
    
    # Setup scheduler for 4:30 PM reminders
    scheduler = BackgroundScheduler()
    scheduler.add_job(
        send_daily_reminders,
        'cron',
        hour=16,  # 4 PM in 24-hour format (adjust for your timezone)
        minute=30,
        args=[app]
    )
    scheduler.start()
    
    logger.info("☕ Coffee Cleaning Reminder Bot started!")
    logger.info("Reminders will be sent daily at 4:30 PM")
    
    # Start polling
    app.run_polling()

if __name__ == '__main__':
    main()
